from django.utils.lorem_ipsum import words
from matplotlib.backends.backend_qt import MainWindow
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from PyQt5 import QtGui
import design
from design import Ui_MainWindow
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QMessageBox, QInputDialog, QLabel, QVBoxLayout, QWidget, QFileDialog
import pandas as pd
import matplotlib.pyplot as plt
import os
from datetime import datetime

class BudgetApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("Бюджетный калькулятор")
        self.ui.label_20.setStyleSheet(u"border-image: url(:/recourceprog/static/icon.png);")
        self.ui.Graficsmall.setStyleSheet(u"border-image: url(:/recourceprog/static/Diagram.png);")
        self.ui.AddSalaryButton.clicked.connect(self.get_income)
        self.ui.Total.clicked.connect(self.display_summary)
        self.ui.AddExpenseButton.clicked.connect(self.add_expense)
        self.ui.BuildGraph.clicked.connect(self.plot_expenses)
        self.ui.TotalSalaryButton.clicked.connect(self.display_total_income)
        self.ui.ClearButton.clicked.connect(self.clear_diagram)
        self.ui.TotalExpensesButton.clicked.connect(self.display_total_expenses)
        self.ui.OpenButton.clicked.connect(self.open_excel_file)
        self.ui.MakeButton.clicked.connect(self.create_excel_file)
        self.expenses = {}
        self.total_income = 0
        self.income_list = []
        self.total_expenses = 0

    def clear_diagram(self):
        self.ui.Grafic.setStyleSheet(u"")

    def get_income(self):
        income_input, ok = QInputDialog.getText(self, "Зарплата", "Введите вашу зарплату:")
        if ok and income_input:
            try:
                self.income = float(income_input)
                self.total_income += self.income
                self.income_list.append(self.income)
                QMessageBox.information(self, "Успех", "Зарплата успешно введена!")
                print(f"Total income {self.total_income}")
            except ValueError:
                QMessageBox.critical(self, "Ошибка", "Пожалуйста, введите корректное число")
                self.get_income()

    def add_expense(self):
        category, ok = QInputDialog.getText(self, "Категория расхода", "Введите категорию расхода:")
        if ok and category:
            amount_input, ok = QInputDialog.getText(self, "Сумма", f"Введите расход на категорию {category}:")
            if ok and amount_input:
                try:
                    amount = float(amount_input)
                    self.expenses[category] = amount
                    self.total_expenses += amount
                    QMessageBox.information(self, "Успех", "Расход успешно добавлен!")
                except ValueError:
                    QMessageBox.critical(self, "Ошибка", "Пожалуйста, введите корректное число")
                    self.add_expense()
        self.plot_expenses()
        print(self.total_expenses)

    def display_summary(self):
        total_expenses = sum(self.expenses.values())
        balance = self.total_income - total_expenses

        summary = (f"Ваша зарплата: {self.total_income}\n"
                   f"Всего трат: {total_expenses}\n"
                   f"Оставшийся баланс: {balance}\n")
        QMessageBox.information(self, "Сводка бюджета", summary)

    def display_total_income(self):
        summary = (f"Всего зарплата: {self.total_income}")
        QMessageBox.information(self, "Сводка бюджета", summary)

    def display_total_expenses(self):
        total_expenses = sum(self.expenses.values())
        summary = (f"Всего зарплата: {self.total_expenses}")
        QMessageBox.information(self, "Сводка бюджета", f"Всего расходов: {self.total_expenses}")

    def plot_expenses(self):
        if not self.expenses:
            QMessageBox.warning(self, "Предупреждение", "Нет расходов для отображения.")
            return

        df = pd.DataFrame(list(self.expenses.items()), columns=["Категория", "Количество"])
        df.plot(kind='bar', x="Категория", y="Количество", legend=True)
        plt.ylabel('Количество')
        plt.title('Диаграмма трат')
        plt.savefig('Diagram.png')
        self.ui.Grafic.setStyleSheet(u"border-image: url(Diagram.png);")

    def open_excel_file(self):
        # Открываем существующий Excel файл
        file_name, _ = QFileDialog.getOpenFileName(self, "Открыть файл", "", "Excel Files (*.xlsx)")
        if file_name:
            workbook = load_workbook(file_name)
            sheet = workbook.active
            # Читаем данные из файла
            income = self.total_income
            expenses = sum(self.expenses.values())
            date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # Добавляем данные в первую свободную строку
            sheet.append([date, income, expenses])
            workbook.save(file_name)
            QMessageBox.information(self, "Успех", "Данные успешно сохранены в Excel файл!")

    def create_excel_file(self):
        # Создаем новый Excel файл
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Бюджет"

        # Добавляем заголовки
        sheet.append(["Дата", "Общий доход", "Общие расходы"])

        # Записываем значения в определенные ячейки
        sheet['A2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Дата в формате YYYY-MM-DD HH:MM:SS
        sheet['B2'] = self.total_income  # Общий доход
        sheet['C2'] = self.total_expenses  # Общие расходы

        # Сохраняем файл с именем
        file_name, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Excel Files (*.xlsx)")
        if file_name:
            workbook.save(file_name)  # Сохраняем файл
            QMessageBox.information(self, "Успех", "Excel файл успешно создан!")

app = QApplication(sys.argv)
budget_app = BudgetApp()
budget_app.show()
sys.exit(app.exec())
